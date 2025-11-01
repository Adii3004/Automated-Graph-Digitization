import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import cv2
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backend_bases import MouseButton
from scipy.optimize import curve_fit
import openpyxl
from openpyxl.styles import PatternFill
import shutil
import os
import customtkinter as ctk

class GraphDigitizerApp:
    def __init__(self, root):
        self.root = root
        self.root.state('zoomed')
        self.root.title("Curve Tracer")

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.root.configure(bg='#000000')

        # Bind the close event to the on_closing method
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Initialize variables
        self.image_path = None
        self.efficiency_image_path = None
        self.img_rgb = None
        self.graph_data = []
        self.clicked_points = []
        self.surge_point = None
        self.collecting = False
        self.fig = None
        self.ax = None
        self.performance_x_interpolated = []
        self.panning = False
        self.cur_xlim = None
        self.cur_ylim = None
        self.design_head = None
        self.design_flowrate = None
        self.design_efficiency = None
        self.rpm_entries = []

        self.pump_template_path = r'C:\Users\shubham gite\OneDrive\Desktop\MIT-WPU\final year\capstone\Final  Curve Tracing App'
        self.compressor_template_path = r'C:\Users\shubham gite\OneDrive\Desktop\MIT-WPU\final year\capstone\Final  Curve Tracing App'
        self.setup_ui()

    def setup_ui(self):
        # Create main frame
        main_frame = ctk.CTkFrame(self.root)
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

        # Image loading
        load_frame = ctk.CTkFrame(main_frame)
        load_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        self.load_button = ctk.CTkButton(load_frame, text="Load Image", command=self.load_image)
        self.load_button.pack(side=tk.LEFT, padx=5, pady=5)

        # Device type selection
        device_frame = ctk.CTkFrame(main_frame)
        device_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=10)
        ctk.CTkLabel(device_frame, text="Device Type").pack(anchor=tk.W)
        self.device_combobox = ctk.CTkComboBox(device_frame, values=["Pump", "Compressor"], state="readonly")
        self.device_combobox.pack(fill=tk.X)
        self.device_combobox.set("Pump")

        # Device details
        details_frame = ctk.CTkFrame(main_frame)
        details_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=10, pady=10)

        ctk.CTkLabel(details_frame, text="Name:").grid(row=0, column=0, sticky=tk.W)
        self.name_entry = ctk.CTkEntry(details_frame)
        self.name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(details_frame, text="Description:").grid(row=1, column=0, sticky=tk.W)
        self.description_entry = ctk.CTkEntry(details_frame)
        self.description_entry.grid(row=1, column=1, sticky=(tk.W, tk.E))

        self.process_efficiency_var = tk.BooleanVar()
        self.process_efficiency_checkbox = ctk.CTkCheckBox(details_frame, text="Process Efficiency Curves", variable=self.process_efficiency_var)
        self.process_efficiency_checkbox.grid(row=2, column=0, columnspan=2, sticky=tk.W)

        ctk.CTkLabel(details_frame, text="Number of Curves:").grid(row=3, column=0, sticky=tk.W)
        self.num_curves_entry = ctk.CTkEntry(details_frame)
        self.num_curves_entry.grid(row=3, column=1, sticky=(tk.W, tk.E))

        # Axis limits and step sizes
        axis_frame = ctk.CTkFrame(main_frame)
        axis_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=10)

        # Performance subheading
        ctk.CTkLabel(axis_frame, text="Performance", font=("Arial", 14)).grid(row=0, column=0, columnspan=2, pady=(0, 5))

        ctk.CTkLabel(axis_frame, text="X Min:").grid(row=1, column=0, sticky=tk.W)
        self.x_min_entry = ctk.CTkEntry(axis_frame)
        self.x_min_entry.grid(row=1, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="X Max:").grid(row=2, column=0, sticky=tk.W)
        self.x_max_entry = ctk.CTkEntry(axis_frame)
        self.x_max_entry.grid(row=2, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Min:").grid(row=3, column=0, sticky=tk.W)
        self.y_min_entry = ctk.CTkEntry(axis_frame)
        self.y_min_entry.grid(row=3, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Max:").grid(row=4, column=0, sticky=tk.W)
        self.y_max_entry = ctk.CTkEntry(axis_frame)
        self.y_max_entry.grid(row=4, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="X Step Size:").grid(row=5, column=0, sticky=tk.W)
        self.x_step_entry = ctk.CTkEntry(axis_frame)
        self.x_step_entry.grid(row=5, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Step Size:").grid(row=6, column=0, sticky=tk.W)
        self.y_step_entry = ctk.CTkEntry(axis_frame)
        self.y_step_entry.grid(row=6, column=1, sticky=(tk.W, tk.E))

        # Efficiency subheading
        ctk.CTkLabel(axis_frame, text="Efficiency", font=("Arial", 14)).grid(row=0, column=2, columnspan=2, pady=(0, 5))

        ctk.CTkLabel(axis_frame, text="X Min:").grid(row=1, column=2, sticky=tk.W)
        self.x_min_eff_entry = ctk.CTkEntry(axis_frame)
        self.x_min_eff_entry.grid(row=1, column=3, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="X Max:").grid(row=2, column=2, sticky=tk.W)
        self.x_max_eff_entry = ctk.CTkEntry(axis_frame)
        self.x_max_eff_entry.grid(row=2, column=3, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Min:").grid(row=3, column=2, sticky=tk.W)
        self.y_min_eff_entry = ctk.CTkEntry(axis_frame)
        self.y_min_eff_entry.grid(row=3, column=3, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Max:").grid(row=4, column=2, sticky=tk.W)
        self.y_max_eff_entry = ctk.CTkEntry(axis_frame)
        self.y_max_eff_entry.grid(row=4, column=3, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="X Step Size:").grid(row=5, column=2, sticky=tk.W)
        self.x_step_eff_entry = ctk.CTkEntry(axis_frame)
        self.x_step_eff_entry.grid(row=5, column=3, sticky=(tk.W, tk.E))

        ctk.CTkLabel(axis_frame, text="Y Step Size:").grid(row=6, column=2, sticky=tk.W)
        self.y_step_eff_entry = ctk.CTkEntry(axis_frame)
        self.y_step_eff_entry.grid(row=6, column=3, sticky=(tk.W, tk.E))

        # Button to copy X-axis values
        self.same_x_button = ctk.CTkButton(axis_frame, text="Same X", command=self.copy_x_values)
        self.same_x_button.grid(row=7, column=0, columnspan=4, pady=10)

        # Design points
        design_frame = ctk.CTkFrame(main_frame)
        design_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=10)

        ctk.CTkLabel(design_frame, text="Design Head:").grid(row=0, column=0, sticky=tk.W)
        self.design_head_entry = ctk.CTkEntry(design_frame)
        self.design_head_entry.grid(row=0, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(design_frame, text="Design Flowrate:").grid(row=1, column=0, sticky=tk.W)
        self.design_flowrate_entry = ctk.CTkEntry(design_frame)
        self.design_flowrate_entry.grid(row=1, column=1, sticky=(tk.W, tk.E))

        ctk.CTkLabel(design_frame, text="Design Efficiency:").grid(row=2, column=0, sticky=tk.W)
        self.design_efficiency_entry = ctk.CTkEntry(design_frame)
        self.design_efficiency_entry.grid(row=2, column=1, sticky=(tk.W, tk.E))

# Adjust the position of the RPM Frame with Scrollbar
        # Adjust the position of the RPM Frame with Scrollbar
        rpm_frame = ctk.CTkFrame(main_frame)
        rpm_frame.grid(row=1, column=2, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

        rpm_label = ctk.CTkLabel(rpm_frame, text="RPM Values:")
        rpm_label.grid(row=0, column=0, columnspan=2, pady=(0, 5), sticky=tk.W)

        # Set the background color directly using a hex value that matches CTkFrame's background
        canvas_bg_color = "#2B2B2B"  # This matches the typical dark mode background of customtkinter

        # Configure the canvas to have the same background as the customtkinter default
        rpm_canvas = tk.Canvas(rpm_frame, height=250, bg=canvas_bg_color)  # Match canvas background to CTkFrame
        rpm_scrollbar = ctk.CTkScrollbar(rpm_frame, orientation="vertical", command=rpm_canvas.yview)
        rpm_canvas.configure(yscrollcommand=rpm_scrollbar.set)

        rpm_canvas.grid(row=1, column=0, sticky=tk.W + tk.E)
        rpm_scrollbar.grid(row=1, column=1, sticky=tk.N + tk.S)

        # Configure the inner frame with the same background color
        rpm_inner_frame = ctk.CTkFrame(rpm_canvas, fg_color=canvas_bg_color)  # Match inner frame background to CTkFrame
        rpm_inner_frame.bind("<Configure>", lambda e: rpm_canvas.configure(scrollregion=rpm_canvas.bbox("all")))

        rpm_canvas.create_window((0, 0), window=rpm_inner_frame, anchor="nw")

        for i in range(12):
            bullet_label = ctk.CTkLabel(rpm_inner_frame, text=f"{i + 1}.", text_color="white")
            bullet_label.grid(row=i, column=0, padx=5, pady=2, sticky=tk.W)
            rpm_entry = ctk.CTkEntry(rpm_inner_frame)
            rpm_entry.grid(row=i, column=1, padx=5, pady=2, sticky=(tk.W, tk.E))
            self.rpm_entries.append(rpm_entry)

        # Make the RPM frame fill its allocated space and look consistent with other frames
        rpm_frame.grid_columnconfigure(0, weight=1)
        rpm_frame.grid_rowconfigure(1, weight=1)
        # Process and undo buttons
        button_frame = ctk.CTkFrame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E))
        self.process_button = ctk.CTkButton(button_frame, text="Process Image", command=self.process_image)
        self.process_button.pack(side=tk.RIGHT, padx=5, pady=5)

        # Configure resizing behavior
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)

    def load_image(self):
        self.image_path = filedialog.askopenfilename()
        if not self.image_path:
            return
        img = cv2.imread(self.image_path)
        if img is None:
            messagebox.showerror("Error", f"Unable to load image '{self.image_path}'. Please check the file path.")
        else:
            self.img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            self.image_path = self.image_path.replace('/', '\\')
            messagebox.showinfo("Success", "Image loaded successfully!")

    def validate_inputs(self):
        try:
            # Retrieve and convert performance values
            self.x_min = float(self.x_min_entry.get())
            self.x_max = float(self.x_max_entry.get())
            self.y_min = float(self.y_min_entry.get())
            self.y_max = float(self.y_max_entry.get())
            self.x_step = float(self.x_step_entry.get())
            self.y_step = float(self.y_step_entry.get())

            # Retrieve and convert efficiency values if processing efficiency curves
            if self.process_efficiency_var.get():
                self.x_min_eff = float(self.x_min_eff_entry.get())
                self.x_max_eff = float(self.x_max_eff_entry.get())
                self.y_min_eff = float(self.y_min_eff_entry.get())
                self.y_max_eff = float(self.y_max_eff_entry.get())
                self.x_step_eff = float(self.x_step_eff_entry.get())
                self.y_step_eff = float(self.y_step_eff_entry.get())

            # Retrieve design values
            self.design_head = float(self.design_head_entry.get())
            self.design_flowrate = float(self.design_flowrate_entry.get())
            self.design_efficiency = float(self.design_efficiency_entry.get())

            # Validation checks
            if not (self.x_max > self.x_min and self.y_max > self.y_min):
                messagebox.showerror("Error", "For Performance: X Max must be greater than X Min and Y Max must be greater than Y Min.")
                return False

            if not (self.x_step < self.x_max and self.y_step < self.y_max):
                messagebox.showerror("Error", "For Performance: X Step must be less than X Max and Y Step must be less than Y Max.")
                return False

            if self.process_efficiency_var.get():
                if not (self.x_max_eff > self.x_min_eff and self.y_max_eff > self.y_min_eff):
                    messagebox.showerror("Error", "For Efficiency: X Max must be greater than X Min and Y Max must be greater than Y Min.")
                    return False

                if not (self.x_step_eff < self.x_max_eff and self.y_step_eff < self.y_max_eff):
                    messagebox.showerror("Error", "For Efficiency: X Step must be less than X Max and Y Step must be less than Y Max.")
                    return False

            # Validation passed
            return True

        except ValueError:
            messagebox.showerror("Error", "Please ensure all numeric fields are filled with valid numbers.")
            return False

    def process_image(self):
        if not self.image_path:
            messagebox.showerror("Error", "Please load an image first.")
            return

        # Ensure validate_inputs is called before accessing attributes
        if not self.validate_inputs():
            return

        device_type = self.device_combobox.get().strip()
        device_name = self.name_entry.get().strip()
        description = self.description_entry.get().strip()

        if not device_type or not device_name or not description:
            messagebox.showerror("Error", "Please enter all required fields.")
            return

        self.filename = f"{device_name}.xlsx" if device_type == "Pump" else f"{device_name}.xlsm"

        self.design_values = {
            'head': self.design_head,
            'flowrate': self.design_flowrate,
            'efficiency': self.design_efficiency
        }

        num_curves = self.num_curves_entry.get()
        if not num_curves.isdigit() or int(num_curves) <= 0:
            messagebox.showerror("Error", "Please enter a valid number of curves.")
            return

        self.num_curves = int(num_curves)

        # Save RPM values to cells A120 to L120
        rpm_values = [entry.get().strip() for entry in self.rpm_entries]
        self.save_rpm_values(rpm_values)

        self.grabit(self.image_path, self.x_min, self.x_max, self.y_min, self.y_max, self.x_step, self.y_step, device_type, device_name, description, efficiency=False)

        if self.process_efficiency_var.get():
            self.load_efficiency_image(self.x_min_eff, self.x_max_eff, self.y_min_eff, self.y_max_eff, device_name, device_type, self.x_step_eff, self.y_step_eff)

    def save_rpm_values(self, rpm_values):
        try:
            if not os.path.exists(self.filename):
                if self.device_combobox.get() == "Pump":
                    shutil.copy(self.pump_template_path, self.filename)
                elif self.device_combobox.get() == "Compressor":
                    shutil.copy(self.compressor_template_path, self.filename)

            workbook = openpyxl.load_workbook(self.filename, keep_vba=True)
            sheet = workbook.active

            # Save RPM values to cells A120 to L120
            for i, value in enumerate(rpm_values):
                cell = f"{chr(65 + i)}120"
                sheet[cell] = value

            workbook.save(self.filename)
        except Exception as e:
            print(f"Error saving RPM values to Excel: {e}")

    def grabit(self, image_path, x_min, x_max, y_min, y_max, x_step, y_step, device_type, device_name, description, efficiency):
        try:
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Error: Unable to load image '{image_path}'. Please check the file path.")

            self.img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

            for curve_index in range(1, self.num_curves + 1):
                self.clicked_points = []
                self.surge_point = None
                self.fig, self.ax = plt.subplots()
                self.undo_button = ttk.Button(self.fig.canvas.manager.window, text="Undo Last Point", command=self.undo_last_point, state=tk.DISABLED)
                self.undo_button.pack(side=tk.RIGHT, padx=5, pady=5)
                self.pan_button = ttk.Button(self.fig.canvas.manager.window, text="Toggle Pan", command=self.toggle_pan)
                self.pan_button.pack(side=tk.RIGHT, padx=5, pady=5)

                def onclick(event):
                    if self.panning:
                        return
                    if event.button == MouseButton.LEFT:
                        print(f"Clicked at ({event.xdata:.2f}, {event.ydata:.2f})")
                        if device_type == "Compressor" and self.surge_point is None:
                            self.surge_point = (event.xdata, event.ydata)
                        else:
                            self.clicked_points.append((event.xdata, event.ydata))
                        self.update_plot()
                    elif event.button == MouseButton.RIGHT:
                        if self.clicked_points:
                            self.clicked_points.pop()
                        elif self.surge_point:
                            self.surge_point = None
                        self.update_plot()

                    if event.dblclick:
                        print("Graph completed")
                        plt.close()

                def on_move(event):
                    if event.inaxes:
                        x, y = event.xdata, event.ydata
                        self.ax.clear()
                        self.ax.imshow(self.img_rgb)
                        self.ax.plot(x, y, 'o', color='red', markersize=3)
                        if self.surge_point:
                            self.ax.plot(self.surge_point[0], self.surge_point[1], 'bo', markersize=3)  # Blue dot for surge point
                        if self.clicked_points:
                            x_coords, y_coords = zip(*self.clicked_points) if self.clicked_points else ([], [])
                            self.ax.plot(x_coords, y_coords, 'ro-', markersize=3)  # Red dots for other points

                        # Apply current zoom limits
                        if self.cur_xlim and self.cur_ylim:
                            self.ax.set_xlim(self.cur_xlim)
                            self.ax.set_ylim(self.cur_ylim)

                        self.fig.canvas.draw()

                self.fig.canvas.mpl_connect('button_press_event', onclick)
                self.fig.canvas.mpl_connect('motion_notify_event', on_move)
                self.fig.canvas.mpl_connect('scroll_event', self.on_scroll)
                self.cidpan = self.fig.canvas.mpl_connect('motion_notify_event', self.on_pan)
                plt.show()

                if not self.clicked_points and self.surge_point is None:
                    print("No points were clicked. Please try again.")
                    continue

                if device_type == "Compressor" and self.surge_point:
                    all_points = [self.surge_point] + self.clicked_points
                else:
                    all_points = self.clicked_points

                points = np.array(all_points)

                x_coords = points[:, 0]
                y_coords = points[:, 1]

                img_height, img_width, _ = self.img_rgb.shape
                x_data = x_min + (x_max - x_min) * (x_coords / img_width)
                y_data = y_max - (y_max - y_min) * (y_coords / img_height)

                self.graph_data.append((x_data, y_data))

                n_interpolated_points = 13
                if efficiency:
                    self.fit_curve(x_data, y_data, n_interpolated_points, device_type, x_min, x_max, self.y_min_eff, self.y_max_eff, self.x_step_eff, self.y_step_eff, device_name, description, curve_index, efficiency)
                else:
                    self.fit_curve(x_data, y_data, n_interpolated_points, device_type, x_min, x_max, y_min, y_max, x_step, y_step, device_name, description, curve_index, efficiency)

            # Print the performance x-interpolated values after all performance curves are processed
            if not efficiency:
                print("Performance x-interpolated values:")
                for i, x_interpolated in enumerate(self.performance_x_interpolated):
                    print(f"Curve {i + 1}: {x_interpolated}")

                if self.process_efficiency_var.get():
                    self.load_efficiency_image(x_min, x_max, y_min, y_max, device_name, device_type, x_step, y_step)
                else:
                    self.open_excel()
            else:
                self.open_excel()

        except Exception as e:
            print(f"Error occurred: {e}")

    def load_efficiency_image(self, x_min, x_max, y_min, y_max, device_name, device_type, x_step, y_step):
        self.efficiency_image_path = filedialog.askopenfilename(title="Load Efficiency Curve Image")
        if self.efficiency_image_path:
            img = cv2.imread(self.efficiency_image_path)
            if img is None:
                messagebox.showerror("Error", f"Unable to load image '{self.efficiency_image_path}'. Please check the file path.")
            else:
                self.img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                self.efficiency_image_path = self.efficiency_image_path.replace('/', '\\')
                self.process_efficiency_image(x_min, x_max, y_min, y_max, device_type, device_name, x_step, y_step)
        else:
            self.open_excel()

    def process_efficiency_image(self, x_min, x_max, y_min, y_max, device_type, device_name, x_step, y_step):
        if not self.efficiency_image_path:
            print("Please load an image first.")
            return

        self.grabit(self.efficiency_image_path, self.x_min_eff, self.x_max_eff, self.y_min_eff, self.y_max_eff, self.x_step_eff, self.y_step_eff, device_type, 'Efficiency', device_name, '', efficiency=True)

    def update_plot(self):
        if self.fig and self.ax:
            # Save current limits to maintain zoom level
            self.cur_xlim = self.ax.get_xlim()
            self.cur_ylim = self.ax.get_ylim()

            self.ax.clear()
            self.ax.imshow(self.img_rgb)
            self.ax.set_title('Graph Image (Left-click to add points, right-click to remove last point, double-click to finish)')
            if self.surge_point:
                self.ax.plot(self.surge_point[0], self.surge_point[1], 'bo', markersize=3)  # Blue dot for surge point
            if self.clicked_points:
                x_coords, y_coords = zip(*self.clicked_points) if self.clicked_points else ([], [])
                self.ax.plot(x_coords, y_coords, 'ro-', markersize=3)  # Red dots for other points

            # Apply zoom values
            if self.cur_xlim and self.cur_ylim:
                self.ax.set_xlim(self.cur_xlim)
                self.ax.set_ylim(self.cur_ylim)

            self.fig.canvas.draw()

            if self.clicked_points or self.surge_point:
                self.undo_button.config(state=tk.NORMAL)
            else:
                self.undo_button.config(state=tk.DISABLED)

    def undo_last_point(self):
        if self.clicked_points:
            self.clicked_points.pop()
        elif self.surge_point:
            self.surge_point = None
        self.update_plot()

        if not self.clicked_points and not self.surge_point:
            self.undo_button.config(state=tk.DISABLED)

    def fit_curve(self, x_data, y_data, n_interpolated_points, device_type, x_min, x_max, y_min, y_max, x_step, y_step, device_name, description, curve_index, efficiency):
        workbook = None
        if not os.path.exists(self.filename):
            if device_type == "Pump":
                shutil.copy(self.pump_template_path, self.filename)
            elif device_type == "Compressor":
                shutil.copy(self.compressor_template_path, self.filename)

        workbook = openpyxl.load_workbook(self.filename, keep_vba=True)

        if efficiency:
            sheet_name = 'Efficiency'
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet["M1"] = x_step
            sheet["N1"] = y_step
        else:
            sheet = workbook.active
            sheet["M1"] = x_step
            sheet["N1"] = y_step
            sheet["C8"] = self.num_curves

        # Save design values if not efficiency sheet
        if not efficiency:
            sheet["B1"] = self.design_values['flowrate']
            sheet["C1"] = self.design_values['head']
            sheet["D1"] = self.design_values['efficiency']

        # Save the image paths
        if not efficiency:
            sheet["A2"] = self.image_path
        else:
            sheet["A3"] = self.efficiency_image_path

        def poly2_func(x, a, b, c):
            return a * x ** 2 + b * x + c

        def poly3_func(x, a, b, c, d):
            return a * x ** 3 + b * x ** 2 + c * x + d

        def poly4_func(x, a, b, c, d, e):
            return a * x ** 4 + b * x ** 3 + c * x ** 2 + d * x + e

        def poly5_func(x, a, b, c, d, e, f):
            return a * x ** 5 + b * x ** 4 + c * x ** 3 + d * x ** 2 + e * x + f

        def poly6_func(x, a, b, c, d, e, f, g):
            return a * x ** 6 + b * x ** 5 + c * x ** 4 + d * x ** 3 + e * x ** 2 + f * x + g

        def calculate_r_squared(y_true, y_pred):
            residuals = y_true - y_pred
            ss_res = np.sum(residuals ** 2)
            ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
            return 1 - (ss_res / ss_tot)

        try:
            fits = [
                ('Polynomial (Order 2)', poly2_func, curve_fit(poly2_func, x_data, y_data)),
                ('Polynomial (Order 3)', poly3_func, curve_fit(poly3_func, x_data, y_data)),
                ('Polynomial (Order 4)', poly4_func, curve_fit(poly4_func, x_data, y_data)),
                ('Polynomial (Order 5)', poly5_func, curve_fit(poly5_func, x_data, y_data)),
                ('Polynomial (Order 6)', poly6_func, curve_fit(poly6_func, x_data, y_data))
            ]
        except RuntimeError as e:
            print(f"RuntimeError: {e}")
            return

        best_fit_name = None
        best_fit_func = None
        best_fit_params = None
        best_r_squared = -np.inf

        for name, func, (params, _) in fits:
            try:
                y_fit = func(x_data, *params)
                r_squared = calculate_r_squared(y_data, y_fit)
                if r_squared > best_r_squared:
                    best_fit_name = name
                    best_fit_func = func
                    best_fit_params = params
                    best_r_squared = r_squared
            except Exception as e:
                print(f"Error in fitting {name}: {e}")
                continue

        first_clicked_x = x_data[0]
        first_clicked_y = y_data[0]
        last_clicked_x = max(x_data)
        if efficiency:
            x_interpolated = self.performance_x_interpolated[curve_index - 1]  # Aligning efficiency curve with respective performance curve
        else:
            x_interpolated = np.linspace(first_clicked_x, last_clicked_x, n_interpolated_points)
            self.performance_x_interpolated.append(x_interpolated)

        y_interpolated = best_fit_func(x_interpolated, *best_fit_params)

        # Ensure y_interpolated values are not negative and set first and last values
        y_interpolated = np.maximum(y_interpolated, 0)

        # Adjust interpolated points before the first clicked point
        if device_type == "Compressor" and self.surge_point:
            surge_x, surge_y = x_data[0], y_data[0]
            distances = np.sqrt((x_interpolated - surge_x) ** 2 + (y_interpolated - surge_y) ** 2)
            closest_index = np.argmin(distances)
            x_interpolated[closest_index] = surge_x
            y_interpolated[closest_index] = surge_y

        if best_fit_name == 'Polynomial (Order 2)':
            equation = f"y = {best_fit_params[0]:.4f}x^2 + {best_fit_params[1]:.4f}x + {best_fit_params[2]:.4f}"
        elif best_fit_name == 'Polynomial (Order 3)':
            equation = f"y = {best_fit_params[0]:.4f}x^3 + {best_fit_params[1]:.4f}x^2 + {best_fit_params[2]:.4f}x + {best_fit_params[3]:.4f}"
        elif best_fit_name == 'Polynomial (Order 4)':
            equation = f"y = {best_fit_params[0]:.4f}x^4 + {best_fit_params[1]:.4f}x^3 + {best_fit_params[2]:.4f}x^2 + {best_fit_params[3]:.4f}x + {best_fit_params[4]:.4f}"
        elif best_fit_name == 'Polynomial (Order 5)':
            equation = f"y = {best_fit_params[0]:.4f}x^5 + {best_fit_params[1]:.4f}x^4 + {best_fit_params[2]:.4f}x^3 + {best_fit_params[3]:.4f}x^2 + {best_fit_params[4]:.4f}x + {best_fit_params[5]:.4f}"
        elif best_fit_name == 'Polynomial (Order 6)':
            equation = f"y = {best_fit_params[0]:.4f}x^6 + {best_fit_params[1]:.4f}x^5 + {best_fit_params[2]:.4f}x^4 + {best_fit_params[3]:.4f}x^3 + {best_fit_params[4]:.4f}x^2 + {best_fit_params[5]:.4f}x + {best_fit_params[6]:.4f}"

        print(f"Best fit: {best_fit_name}")
        print(f"Equation: {equation}")
        print(f"R-squared: {best_r_squared:.4f}")

        print("Interpolated points:")
        for x, y in zip(x_interpolated, y_interpolated):
            print(f"x: {x:.4f}, y: {y:.4f}")

        if device_type == "Pump":
            sheet_name = device_name if not efficiency else 'Efficiency'
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                template_sheet = workbook.active
                sheet = workbook.copy_worksheet(template_sheet)
                sheet.title = sheet_name
            if efficiency:
                self.save_to_excel(sheet, x_interpolated, y_interpolated, "C14", "D14")
            else:
                self.save_to_excel(sheet, x_interpolated, y_interpolated, "C8", "D8")
        elif device_type == "Compressor":
            col_offset = (curve_index - 1) * 4
            x_start_col = self.get_column_letter(2 + col_offset)  # Adjust column index properly
            y_start_col = self.get_column_letter(3 + col_offset)
            x_start = f"{x_start_col}14"
            y_start = f"{y_start_col}14"
            self.save_to_excel(sheet, x_interpolated, y_interpolated, x_start, y_start, surge_index=closest_index if self.surge_point else None)

            surge_x_start = f"V{33 + (curve_index - 1)}"
            surge_y_start = f"W{33 + (curve_index - 1)}"
            if self.surge_point:
                surge_x_data = x_min + (x_max - x_min) * (self.surge_point[0] / self.img_rgb.shape[1])
                surge_y_data = y_max - (y_max - y_min) * (self.surge_point[1] / self.img_rgb.shape[0])
                self.save_surge_points(sheet, surge_x_data, surge_y_data, surge_x_start, surge_y_start)

        sheet["C4"] = x_min
        sheet["C5"] = x_max
        sheet["C6"] = y_min
        sheet["C7"] = y_max

        if efficiency:
            sheet["C10"] = self.y_min_eff
            sheet["C11"] = self.y_max_eff

        sheet["C201"] = device_name
        sheet["C202"] = description

        curve_col = 2 + (curve_index - 1) * 4
        curve_cell = f"{openpyxl.utils.get_column_letter(curve_col)}200"
        sheet[curve_cell] = description

        workbook.save(self.filename)

    def get_column_letter(self, n):
        """ Convert a number to an Excel column letter (1 -> A, 27 -> AA). """
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def save_to_excel(self, sheet, x_interpolated, y_interpolated, x_cell_start, y_cell_start, surge_index=None):
        try:
            # Extract the starting row and column from the cell references
            start_row = int(''.join(filter(str.isdigit, x_cell_start)))
            start_col_num = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, x_cell_start)))

            for i, (x, y) in enumerate(zip(x_interpolated, y_interpolated)):
                x_cell = f"{self.get_column_letter(start_col_num)}{start_row + i}"
                y_cell = f"{self.get_column_letter(start_col_num + 1)}{start_row + i}"
                sheet[x_cell] = x
                sheet[y_cell] = y

                if surge_index is not None and i == surge_index:
                    fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
                    sheet[x_cell].fill = fill
                    sheet[y_cell].fill = fill
        except Exception as e:
            print(f"Error saving to Excel: {e}")

    def save_surge_points(self, sheet, surge_x, surge_y, x_start, y_start):
        sheet[x_start] = surge_x
        sheet[y_start] = surge_y
        fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
        sheet[x_start].fill = fill
        sheet[y_start].fill = fill

    def display_fitted_curve(self, x_data, y_data, x_interpolated, y_interpolated, best_fit_name):
        self.fig, self.ax = plt.subplots()
        self.ax.imshow(self.img_rgb)
        self.ax.plot(x_data, y_data, 'ro', label='Data')
        self.ax.plot(x_interpolated, y_interpolated, 'b-', label=f'Best fit ({best_fit_name})')
        self.ax.legend()
        self.ax.set_title('Graph Image with Fitted Curve (Left-click to add points, right-click to remove last point, double-click to finish)')
        self.fig.canvas.mpl_connect('button_press_event', self.on_click_adjust)
        plt.show()

    def on_click_adjust(self, event):
        if event.button == MouseButton.LEFT:
            self.clicked_points.append((event.xdata, event.ydata))
        elif event.button == MouseButton.RIGHT and self.clicked_points:
            self.clicked_points.pop()
        self.update_plot()

    def open_excel(self):
        os.startfile(self.filename)

    def on_scroll(self, event):
        base_scale = 1.1
        self.cur_xlim = self.ax.get_xlim()
        self.cur_ylim = self.ax.get_ylim()
        xdata = event.xdata
        ydata = event.ydata

        if event.button == 'up':
            scale_factor = 1 / base_scale
        elif event.button == 'down':
            scale_factor = base_scale
        else:
            scale_factor = 1

        new_width = (self.cur_xlim[1] - self.cur_xlim[0]) * scale_factor
        new_height = (self.cur_ylim[1] - self.cur_ylim[0]) * scale_factor
        relx = (xdata - self.cur_xlim[0]) / (self.cur_xlim[1] - self.cur_xlim[0])
        rely = (ydata - self.cur_ylim[0]) / (self.cur_ylim[1] - self.cur_ylim[0])

        self.ax.set_xlim([xdata - new_width * relx, xdata + new_width * (1 - relx)])
        self.ax.set_ylim([ydata - new_height * rely, ydata + new_height * (1 - rely)])
        self.fig.canvas.draw()

    def toggle_pan(self):
        self.panning = not self.panning
        if self.panning:
            self.cidpan = self.fig.canvas.mpl_connect('motion_notify_event', self.on_pan)
            self.pan_button.config(text="Disable Pan")
        else:
            self.fig.canvas.mpl_disconnect(self.cidpan)
            self.pan_button.config(text="Enable Pan")

    def on_pan(self, event):
        if not self.panning:
            return
        if event.button == MouseButton.LEFT:
            dx = event.x - self.prev_event.x
            dy = event.y - self.prev_event.y
            self.cur_xlim = self.ax.get_xlim()
            self.cur_ylim = self.ax.get_ylim()
            new_xlim = [self.cur_xlim[0] - dx, self.cur_xlim[1] - dx]
            new_ylim = [self.cur_ylim[0] - dy, self.cur_ylim[1] - dy]
            self.ax.set_xlim(new_xlim)
            self.ax.set_ylim(new_ylim)
            self.fig.canvas.draw()
        self.prev_event = event

    def copy_x_values(self):
        """Copy X Min, X Max, and X Step Size between performance and efficiency."""
        # Determine which set of fields are filled in first
        if self.x_min_entry.get() and self.x_max_entry.get() and self.x_step_entry.get():
            # Copy from performance to efficiency
            self.x_min_eff_entry.delete(0, tk.END)
            self.x_min_eff_entry.insert(0, self.x_min_entry.get())

            self.x_max_eff_entry.delete(0, tk.END)
            self.x_max_eff_entry.insert(0, self.x_max_entry.get())

            self.x_step_eff_entry.delete(0, tk.END)
            self.x_step_eff_entry.insert(0, self.x_step_entry.get())

        elif self.x_min_eff_entry.get() and self.x_max_eff_entry.get() and self.x_step_eff_entry.get():
            # Copy from efficiency to performance
            self.x_min_entry.delete(0, tk.END)
            self.x_min_entry.insert(0, self.x_min_eff_entry.get())

            self.x_max_entry.delete(0, tk.END)
            self.x_max_entry.insert(0, self.x_max_eff_entry.get())

            self.x_step_entry.delete(0, tk.END)
            self.x_step_entry.insert(0, self.x_step_eff_entry.get())
        else:
            messagebox.showinfo("Info", "Please fill in either performance or efficiency fields first.")

    def on_closing(self):
        """Handle the close event of the main window."""
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.root.destroy()
            os._exit(0)  # Forcefully terminate the application


if __name__ == "__main__":
    root = tk.Tk()
    app = GraphDigitizerApp(root)
    root.mainloop()